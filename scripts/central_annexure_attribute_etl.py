"""
Central, Annexure, and Attribute Reports ETL Pipeline
=====================================================
Extracts data from:
1. Central Reports (1 sheet: GEC) - District-level groundwater assessment
2. Annexure 1 (1 sheet: Annexure 1A) - State-level summary
3. Annexure 2 (1 sheet: Annexure II) - District-level groundwater resources by state
4. Annexure 3 (6 sheets: 3A-3F) - Categorization data
5. Annexure 4 (2 sheets: 4A, 4B) - Assessment unit categorization & quality problems
6. Attribute Reports (2 sheets: Summary, Table) - Detailed assessment unit data

Output: Individual JSON files + Master JSON combining all data
"""

import openpyxl
import json
import os
import re
from pathlib import Path
from datetime import datetime

DATA_DIR = Path(__file__).parent.parent / "data"
OUTPUT_DIR = Path(__file__).parent.parent / "output"


def safe_float(value, default=0.0):
    """Safely convert value to float, return default if conversion fails"""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).strip().replace(",", "")
        if cleaned in ("", "-", "NA", "N/A", "NR", "--"):
            return default
        return float(cleaned)
    except (ValueError, TypeError):
        return default


def safe_str(value, default=""):
    """Safely convert value to string"""
    if value is None:
        return default
    return str(value).strip()


def clean_column_name(name):
    """Clean column name for JSON key"""
    if not name:
        return ""
    name = str(name).strip()
    # Remove special characters and normalize
    name = re.sub(r"[^\w\s]", "", name)
    name = re.sub(r"\s+", "_", name)
    name = name.lower()
    return name


def extract_year_from_filename(filename):
    """Extract year from filename like CentralReport2021-22.xlsx"""
    match = re.search(r"(\d{4})[-_]?(\d{2,4})", filename)
    if match:
        year1 = match.group(1)
        year2 = match.group(2)
        if len(year2) == 2:
            year2 = "20" + year2
        return f"{year1}-{year2}"
    return "unknown"


# =============================================================================
# CENTRAL REPORT ETL
# =============================================================================


def extract_central_report(filepath):
    """Extract Central Report data (GEC sheet)"""
    print(f"  Processing: {filepath.name}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["GEC"]
    year = extract_year_from_filename(filepath.name)

    records = []

    # Data starts from row 12, headers are in rows 8-10
    # Structure similar to State Reports but at district level
    for row_idx in range(12, ws.max_row + 1):
        # Check if row has data (column 2 = STATE should not be empty for data rows)
        state = ws.cell(row=row_idx, column=2).value
        if not state or str(state).strip() in ("", "Total", "TOTAL"):
            continue

        record = {
            "serial_no": safe_float(ws.cell(row=row_idx, column=1).value),
            "state": safe_str(state),
            "district": safe_str(ws.cell(row=row_idx, column=3).value),
            "assessment_unit": safe_str(ws.cell(row=row_idx, column=4).value),
            "year": year,
            # Rainfall (mm) - columns 5-8
            "rainfall_mm": {
                "contaminated": safe_float(ws.cell(row=row_idx, column=5).value),
                "non_contaminated": safe_float(ws.cell(row=row_idx, column=6).value),
                "poor_quality": safe_float(ws.cell(row=row_idx, column=7).value),
                "total": safe_float(ws.cell(row=row_idx, column=8).value),
            },
            # Total Geographical Area - columns 9-14
            "total_geographical_area_ha": {
                "recharge_worthy_area": {
                    "contaminated": safe_float(ws.cell(row=row_idx, column=9).value),
                    "non_contaminated": safe_float(
                        ws.cell(row=row_idx, column=10).value
                    ),
                    "poor_quality": safe_float(ws.cell(row=row_idx, column=11).value),
                    "total": safe_float(ws.cell(row=row_idx, column=12).value),
                },
                "hilly_area": safe_float(ws.cell(row=row_idx, column=13).value),
                "total": safe_float(ws.cell(row=row_idx, column=14).value),
            },
            # Ground Water Recharge - columns 15-60 (similar structure to state reports)
            "ground_water_recharge_ham": {
                "rainfall_recharge": {
                    "contaminated": safe_float(ws.cell(row=row_idx, column=15).value),
                    "non_contaminated": safe_float(
                        ws.cell(row=row_idx, column=16).value
                    ),
                    "poor_quality": safe_float(ws.cell(row=row_idx, column=17).value),
                    "total": safe_float(ws.cell(row=row_idx, column=18).value),
                },
                "canals": {
                    "contaminated": safe_float(ws.cell(row=row_idx, column=19).value),
                    "non_contaminated": safe_float(
                        ws.cell(row=row_idx, column=20).value
                    ),
                    "poor_quality": safe_float(ws.cell(row=row_idx, column=21).value),
                    "total": safe_float(ws.cell(row=row_idx, column=22).value),
                },
                "tanks_ponds": {
                    "contaminated": safe_float(ws.cell(row=row_idx, column=23).value),
                    "non_contaminated": safe_float(
                        ws.cell(row=row_idx, column=24).value
                    ),
                    "poor_quality": safe_float(ws.cell(row=row_idx, column=25).value),
                    "total": safe_float(ws.cell(row=row_idx, column=26).value),
                },
                "water_conservation_structures": {
                    "contaminated": safe_float(ws.cell(row=row_idx, column=27).value),
                    "non_contaminated": safe_float(
                        ws.cell(row=row_idx, column=28).value
                    ),
                    "poor_quality": safe_float(ws.cell(row=row_idx, column=29).value),
                    "total": safe_float(ws.cell(row=row_idx, column=30).value),
                },
            },
            # Additional columns - simplified extraction for central report
            "recharge_from_other_sources_monsoon_ham": safe_float(
                ws.cell(row=row_idx, column=31).value
            ),
            "recharge_from_rainfall_non_monsoon_ham": safe_float(
                ws.cell(row=row_idx, column=35).value
            ),
            "total_annual_recharge_ham": safe_float(
                ws.cell(row=row_idx, column=51).value
            ),
            "natural_discharge_ham": safe_float(ws.cell(row=row_idx, column=55).value),
            "annual_extractable_resource_ham": safe_float(
                ws.cell(row=row_idx, column=59).value
            ),
            # Extraction data
            "current_extraction_ham": {
                "irrigation": safe_float(ws.cell(row=row_idx, column=63).value),
                "domestic": safe_float(ws.cell(row=row_idx, column=67).value),
                "industrial": safe_float(ws.cell(row=row_idx, column=71).value),
                "total": safe_float(ws.cell(row=row_idx, column=75).value),
            },
            # Stage and categorization
            "stage_of_extraction_percent": safe_float(
                ws.cell(row=row_idx, column=150).value
            ),
            "categorization": safe_str(ws.cell(row=row_idx, column=151).value),
        }

        records.append(record)

    wb.close()
    print(f"    → Extracted {len(records)} district records")
    return year, records


# =============================================================================
# ANNEXURE 1 ETL (State-level Summary)
# =============================================================================


def extract_annexure1(filepath):
    """Extract Annexure 1A - State-level groundwater summary"""
    print(f"  Processing: {filepath.name}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Annexure 1A"]
    year = extract_year_from_filename(filepath.name)

    records = []

    # Headers in rows 3-5, data starts from row 7
    for row_idx in range(7, ws.max_row + 1):
        state = ws.cell(row=row_idx, column=2).value
        if not state or str(state).strip() in (
            "",
            "Total",
            "TOTAL",
            "All India",
            "ALL INDIA",
        ):
            continue

        record = {
            "serial_no": safe_float(ws.cell(row=row_idx, column=1).value),
            "state": safe_str(state),
            "year": year,
            # Ground Water Recharge (BCM)
            "ground_water_recharge_bcm": {
                "monsoon_season": {
                    "from_rainfall": safe_float(ws.cell(row=row_idx, column=3).value),
                    "from_other_sources": safe_float(
                        ws.cell(row=row_idx, column=4).value
                    ),
                },
                "non_monsoon_season": {
                    "from_rainfall": safe_float(ws.cell(row=row_idx, column=5).value),
                    "from_other_sources": safe_float(
                        ws.cell(row=row_idx, column=6).value
                    ),
                },
                "total_annual": safe_float(ws.cell(row=row_idx, column=7).value),
            },
            "total_natural_discharges_bcm": safe_float(
                ws.cell(row=row_idx, column=8).value
            ),
            "annual_extractable_resource_bcm": safe_float(
                ws.cell(row=row_idx, column=9).value
            ),
            # Current Annual Extraction
            "current_annual_extraction_bcm": {
                "irrigation": safe_float(ws.cell(row=row_idx, column=10).value),
                "industrial": safe_float(ws.cell(row=row_idx, column=11).value),
                "domestic": safe_float(ws.cell(row=row_idx, column=12).value),
                "total": safe_float(ws.cell(row=row_idx, column=13).value),
            },
            "annual_gw_allocation_domestic_bcm": safe_float(
                ws.cell(row=row_idx, column=14).value
            ),
            "net_gw_availability_bcm": safe_float(
                ws.cell(row=row_idx, column=15).value
            ),
            "stage_of_extraction_percent": safe_float(
                ws.cell(row=row_idx, column=16).value
            ),
        }

        records.append(record)

    wb.close()
    print(f"    → Extracted {len(records)} state records")
    return year, records


# =============================================================================
# ANNEXURE 2 ETL (District-level by State)
# =============================================================================


def extract_annexure2(filepath):
    """Extract Annexure II - District-level data organized by state"""
    print(f"  Processing: {filepath.name}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Annexure II"]
    year = extract_year_from_filename(filepath.name)

    records = []
    current_state = None

    # Data is organized by state sections with headers repeated
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value

        # Check for state header (row with state name in column 1, "DYNAMIC" or state name alone)
        if cell_a and "DYNAMIC GROUND WATER" in str(cell_a).upper():
            # Next row has state name
            next_row_val = ws.cell(row=row_idx + 1, column=1).value
            if next_row_val:
                current_state = safe_str(next_row_val)
            continue

        # Skip header rows and totals
        if not cell_a or str(cell_a).strip() in (
            "",
            "S.NO",
            "S.No",
            "1",
            "Total(Ham)",
            "Total(Bcm)",
        ):
            # Check if it's a data row with serial number
            try:
                serial = float(cell_a) if cell_a else None
                if serial is None or serial < 1:
                    continue
            except (ValueError, TypeError):
                continue

        district = ws.cell(row=row_idx, column=2).value
        if not district or str(district).strip() in (
            "",
            "Name of District",
            "Total",
            "Total(Ham)",
            "Total(Bcm)",
        ):
            continue

        record = {
            "serial_no": safe_float(ws.cell(row=row_idx, column=1).value),
            "state": current_state or "Unknown",
            "district": safe_str(district),
            "year": year,
            # Ground Water Recharge (HAM)
            "ground_water_recharge_ham": {
                "monsoon_season": {
                    "from_rainfall": safe_float(ws.cell(row=row_idx, column=3).value),
                    "from_other_sources": safe_float(
                        ws.cell(row=row_idx, column=4).value
                    ),
                },
                "non_monsoon_season": {
                    "from_rainfall": safe_float(ws.cell(row=row_idx, column=5).value),
                    "from_other_sources": safe_float(
                        ws.cell(row=row_idx, column=6).value
                    ),
                },
                "total_annual": safe_float(ws.cell(row=row_idx, column=7).value),
            },
            "total_natural_discharges_ham": safe_float(
                ws.cell(row=row_idx, column=8).value
            ),
            "annual_extractable_resource_ham": safe_float(
                ws.cell(row=row_idx, column=9).value
            ),
            "current_annual_extraction_ham": {
                "irrigation": safe_float(ws.cell(row=row_idx, column=10).value),
                "industrial": safe_float(ws.cell(row=row_idx, column=11).value),
                "domestic": safe_float(ws.cell(row=row_idx, column=12).value),
                "total": safe_float(ws.cell(row=row_idx, column=13).value),
            },
            "annual_gw_allocation_domestic_ham": safe_float(
                ws.cell(row=row_idx, column=14).value
            ),
            "net_gw_availability_ham": safe_float(
                ws.cell(row=row_idx, column=15).value
            ),
            "stage_of_extraction_percent": safe_float(
                ws.cell(row=row_idx, column=16).value
            ),
        }

        records.append(record)

    wb.close()
    print(f"    → Extracted {len(records)} district records")
    return year, records


# =============================================================================
# ANNEXURE 3 ETL (Multiple Sheets - Categorization Data)
# =============================================================================


def extract_annexure3(filepath):
    """Extract Annexure 3 with all 6 sheets (3A-3F)"""
    print(f"  Processing: {filepath.name}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    year = extract_year_from_filename(filepath.name)

    all_data = {
        "year": year,
        "annexure_3a_state_categorization": [],
        "annexure_3b_district_categorization": [],
        "annexure_3c_state_extractable_resource": [],
        "annexure_3d_district_extractable_resource": [],
        "annexure_3e_state_area_categorization": [],
        "annexure_3f_district_area_categorization": [],
    }

    # Sheet 3A - State-level categorization (number of units)
    if "Annexure-3A" in wb.sheetnames:
        ws = wb["Annexure-3A"]
        for row_idx in range(5, ws.max_row + 1):
            state = ws.cell(row=row_idx, column=2).value
            if not state or str(state).strip() in ("", "States", "Total", "ALL INDIA"):
                continue

            record = {
                "serial_no": safe_float(ws.cell(row=row_idx, column=1).value),
                "state": safe_str(state),
                "total_assessed_units": safe_float(
                    ws.cell(row=row_idx, column=3).value
                ),
                "safe": {
                    "count": safe_float(ws.cell(row=row_idx, column=4).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=5).value),
                },
                "semi_critical": {
                    "count": safe_float(ws.cell(row=row_idx, column=6).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=7).value),
                },
                "critical": {
                    "count": safe_float(ws.cell(row=row_idx, column=8).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=9).value),
                },
                "over_exploited": {
                    "count": safe_float(ws.cell(row=row_idx, column=10).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=11).value),
                },
                "saline": {
                    "count": safe_float(ws.cell(row=row_idx, column=12).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=13).value),
                },
            }
            all_data["annexure_3a_state_categorization"].append(record)

    # Sheet 3B - District-level categorization (organized by state)
    if "Annexure-3B" in wb.sheetnames:
        ws = wb["Annexure-3B"]
        current_state = None

        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value

            if cell_a and "DYNAMIC GROUND WATER" in str(cell_a).upper():
                next_val = ws.cell(row=row_idx + 1, column=1).value
                if next_val and str(next_val).strip() not in ("S.No", "S.NO"):
                    current_state = safe_str(next_val)
                continue

            district = ws.cell(row=row_idx, column=2).value
            if not district or str(district).strip() in (
                "",
                "Name of District",
                "Total",
                "No",
                "No.",
            ):
                continue

            try:
                serial = float(cell_a) if cell_a else None
                if serial is None or serial < 1:
                    continue
            except (ValueError, TypeError):
                continue

            record = {
                "serial_no": safe_float(cell_a),
                "state": current_state or "Unknown",
                "district": safe_str(district),
                "total_assessed_units": safe_float(
                    ws.cell(row=row_idx, column=3).value
                ),
                "safe": {
                    "count": safe_float(ws.cell(row=row_idx, column=4).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=5).value),
                },
                "semi_critical": {
                    "count": safe_float(ws.cell(row=row_idx, column=6).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=7).value),
                },
                "critical": {
                    "count": safe_float(ws.cell(row=row_idx, column=8).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=9).value),
                },
                "over_exploited": {
                    "count": safe_float(ws.cell(row=row_idx, column=10).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=11).value),
                },
                "saline": {
                    "count": safe_float(ws.cell(row=row_idx, column=12).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=13).value),
                },
            }
            all_data["annexure_3b_district_categorization"].append(record)

    # Sheet 3C - State-level extractable resource by category
    if "Annexure-3C" in wb.sheetnames:
        ws = wb["Annexure-3C"]
        for row_idx in range(6, ws.max_row + 1):
            state = ws.cell(row=row_idx, column=2).value
            if not state or str(state).strip() in ("", "State", "Total", "ALL INDIA"):
                continue

            record = {
                "serial_no": safe_float(ws.cell(row=row_idx, column=1).value),
                "state": safe_str(state),
                "total_extractable_resource_ham": safe_float(
                    ws.cell(row=row_idx, column=3).value
                ),
                "safe": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=4).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=5).value),
                },
                "semi_critical": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=6).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=7).value),
                },
                "critical": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=8).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=9).value),
                },
                "over_exploited": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=10).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=11).value),
                },
                "saline": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=12).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=13).value),
                },
            }
            all_data["annexure_3c_state_extractable_resource"].append(record)

    # Sheet 3D - District-level extractable resource by category
    if "Annexure-3D" in wb.sheetnames:
        ws = wb["Annexure-3D"]
        current_state = None

        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value

            if cell_a and "DYNAMIC GROUND WATER" in str(cell_a).upper():
                next_val = ws.cell(row=row_idx + 1, column=1).value
                if next_val and str(next_val).strip() not in ("S.No", "S.NO"):
                    current_state = safe_str(next_val)
                continue

            district = ws.cell(row=row_idx, column=2).value
            if not district or str(district).strip() in (
                "",
                "Name of District",
                "Total",
            ):
                continue

            try:
                serial = float(cell_a) if cell_a else None
                if serial is None or serial < 1:
                    continue
            except (ValueError, TypeError):
                continue

            record = {
                "serial_no": safe_float(cell_a),
                "state": current_state or "Unknown",
                "district": safe_str(district),
                "total_extractable_resource_ham": safe_float(
                    ws.cell(row=row_idx, column=3).value
                ),
                "safe": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=4).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=5).value),
                },
                "semi_critical": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=6).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=7).value),
                },
                "critical": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=8).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=9).value),
                },
                "over_exploited": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=10).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=11).value),
                },
                "saline": {
                    "resource_ham": safe_float(ws.cell(row=row_idx, column=12).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=13).value),
                },
            }
            all_data["annexure_3d_district_extractable_resource"].append(record)

    # Sheet 3E - State-level area by category
    if "Annexure-3E" in wb.sheetnames:
        ws = wb["Annexure-3E"]
        for row_idx in range(5, ws.max_row + 1):
            state = ws.cell(row=row_idx, column=2).value
            if not state or str(state).strip() in ("", "States", "Total", "ALL INDIA"):
                continue

            record = {
                "serial_no": safe_float(ws.cell(row=row_idx, column=1).value),
                "state": safe_str(state),
                "total_geographical_area_sq_km": safe_float(
                    ws.cell(row=row_idx, column=3).value
                ),
                "recharge_worthy_area_sq_km": safe_float(
                    ws.cell(row=row_idx, column=4).value
                ),
                "safe": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=5).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=6).value),
                },
                "semi_critical": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=7).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=8).value),
                },
                "critical": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=9).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=10).value),
                },
                "over_exploited": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=11).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=12).value),
                },
                "saline": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=13).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=14).value),
                },
            }
            all_data["annexure_3e_state_area_categorization"].append(record)

    # Sheet 3F - District-level area by category
    if "Annexure-3F" in wb.sheetnames:
        ws = wb["Annexure-3F"]
        current_state = None

        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value

            if cell_a and "DYNAMIC GROUND WATER" in str(cell_a).upper():
                next_val = ws.cell(row=row_idx + 1, column=1).value
                if next_val and str(next_val).strip() not in ("S.No", "S.NO"):
                    current_state = safe_str(next_val)
                continue

            district = ws.cell(row=row_idx, column=2).value
            if not district or str(district).strip() in (
                "",
                "Name of District",
                "Total",
            ):
                continue

            try:
                serial = float(cell_a) if cell_a else None
                if serial is None or serial < 1:
                    continue
            except (ValueError, TypeError):
                continue

            record = {
                "serial_no": safe_float(cell_a),
                "state": current_state or "Unknown",
                "district": safe_str(district),
                "total_recharge_worthy_area_sq_km": safe_float(
                    ws.cell(row=row_idx, column=3).value
                ),
                "safe": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=4).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=5).value),
                },
                "semi_critical": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=6).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=7).value),
                },
                "critical": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=8).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=9).value),
                },
                "over_exploited": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=10).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=11).value),
                },
                "saline": {
                    "area_sq_km": safe_float(ws.cell(row=row_idx, column=12).value),
                    "percent": safe_float(ws.cell(row=row_idx, column=13).value),
                },
            }
            all_data["annexure_3f_district_area_categorization"].append(record)

    wb.close()

    total_records = sum(len(v) for k, v in all_data.items() if isinstance(v, list))
    print(f"    → Extracted {total_records} records across 6 sheets")

    return year, all_data


# =============================================================================
# ANNEXURE 4 ETL (Assessment Unit Categorization & Quality Problems)
# =============================================================================


def extract_annexure4(filepath):
    """Extract Annexure 4 with sheets 4A and 4B"""
    print(f"  Processing: {filepath.name}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    year = extract_year_from_filename(filepath.name)

    all_data = {
        "year": year,
        "annexure_4a_categorized_units": [],  # Semi-critical, Critical, Over-exploited names
        "annexure_4b_quality_problems": [],  # Units with quality issues
    }

    # Sheet 4A - Categorization of assessment units (names by district)
    if "Annexure 4A" in wb.sheetnames:
        ws = wb["Annexure 4A"]
        current_state = None

        for row_idx in range(2, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value

            # Check for state section header
            if cell_a and "CATEGORISATION OF ASSESSMENT" in str(cell_a).upper():
                next_val = ws.cell(row=row_idx + 1, column=1).value
                if next_val and str(next_val).strip() not in (
                    "S.NO",
                    "S.No",
                    "ABSTRACT",
                ):
                    current_state = safe_str(next_val)
                continue

            district = ws.cell(row=row_idx, column=2).value
            if not district or str(district).strip() in (
                "",
                "Name of District",
                "ABSTRACT",
                "Total No. of Assessed Units",
            ):
                continue

            try:
                serial = float(cell_a) if cell_a else None
                if serial is None or serial < 1:
                    continue
            except (ValueError, TypeError):
                continue

            # Extract unit names from columns
            semi_critical_name = safe_str(ws.cell(row=row_idx, column=4).value)
            critical_name = safe_str(ws.cell(row=row_idx, column=6).value)
            over_exploited_name = safe_str(ws.cell(row=row_idx, column=8).value)

            if semi_critical_name or critical_name or over_exploited_name:
                record = {
                    "state": current_state or "Unknown",
                    "district": safe_str(district),
                    "semi_critical_unit": semi_critical_name,
                    "critical_unit": critical_name,
                    "over_exploited_unit": over_exploited_name,
                }
                all_data["annexure_4a_categorized_units"].append(record)

    # Sheet 4B - Quality problems in assessment units
    if "Annexure 4B" in wb.sheetnames:
        ws = wb["Annexure 4B"]
        current_state = None

        for row_idx in range(2, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value

            if cell_a and "QUALITY PROBLEMS" in str(cell_a).upper():
                next_val = ws.cell(row=row_idx + 1, column=1).value
                if next_val and str(next_val).strip() not in (
                    "S.NO",
                    "S.No",
                    "ABSTRACT",
                ):
                    current_state = safe_str(next_val)
                continue

            district = ws.cell(row=row_idx, column=2).value
            if not district or str(district).strip() in (
                "",
                "Name of District",
                "ABSTRACT",
                "Total No. of Assessed Units",
            ):
                continue

            try:
                serial = float(cell_a) if cell_a else None
                if serial is None or serial < 1:
                    continue
            except (ValueError, TypeError):
                continue

            # Quality problem unit names (columns 4, 6, 8 typically)
            fluoride_affected = safe_str(ws.cell(row=row_idx, column=4).value)
            arsenic_affected = safe_str(ws.cell(row=row_idx, column=6).value)
            salinity_affected = safe_str(ws.cell(row=row_idx, column=8).value)

            if fluoride_affected or arsenic_affected or salinity_affected:
                record = {
                    "state": current_state or "Unknown",
                    "district": safe_str(district),
                    "fluoride_affected_unit": fluoride_affected,
                    "arsenic_affected_unit": arsenic_affected,
                    "salinity_affected_unit": salinity_affected,
                }
                all_data["annexure_4b_quality_problems"].append(record)

    wb.close()

    total_records = len(all_data["annexure_4a_categorized_units"]) + len(
        all_data["annexure_4b_quality_problems"]
    )
    print(f"    → Extracted {total_records} records across 2 sheets")

    return year, all_data


# =============================================================================
# ATTRIBUTE REPORT ETL
# =============================================================================


def extract_attribute_report(filepath):
    """Extract Attribute Report with Summary and Table sheets"""
    print(f"  Processing: {filepath.name}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    year = extract_year_from_filename(filepath.name)

    all_data = {"year": year, "summary": [], "detailed_table": []}

    # Sheet: Summary - State-level categorization summary
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        for row_idx in range(5, ws.max_row + 1):
            state = ws.cell(row=row_idx, column=1).value
            if not state or str(state).strip() in ("", "State", "Total", "Grand Total"):
                continue

            record = {
                "state": safe_str(state),
                "over_exploited_count": safe_float(
                    ws.cell(row=row_idx, column=2).value
                ),
                "safe_count": safe_float(ws.cell(row=row_idx, column=3).value),
                "saline_count": safe_float(ws.cell(row=row_idx, column=4).value),
                "critical_count": safe_float(ws.cell(row=row_idx, column=5).value),
                "semi_critical_count": safe_float(ws.cell(row=row_idx, column=6).value),
                "total_count": safe_float(ws.cell(row=row_idx, column=7).value),
            }
            all_data["summary"].append(record)

    # Sheet: Table - Detailed assessment unit data (26 columns)
    if "Table" in wb.sheetnames:
        ws = wb["Table"]

        for row_idx in range(3, ws.max_row + 1):
            state = ws.cell(row=row_idx, column=5).value
            if not state or str(state).strip() == "":
                continue

            record = {
                "serial_no": safe_float(ws.cell(row=row_idx, column=1).value),
                "state_code": safe_str(ws.cell(row=row_idx, column=2).value),
                "state_district_code": safe_str(ws.cell(row=row_idx, column=3).value),
                "state_district_block_code": safe_str(
                    ws.cell(row=row_idx, column=4).value
                ),
                "state": safe_str(state),
                "district": safe_str(ws.cell(row=row_idx, column=6).value),
                "assessment_unit_name": safe_str(ws.cell(row=row_idx, column=7).value),
                "assessment_unit_type": safe_str(ws.cell(row=row_idx, column=8).value),
                "total_geographical_area_ha": safe_float(
                    ws.cell(row=row_idx, column=9).value
                ),
                "recharge_worthy_area_ha": safe_float(
                    ws.cell(row=row_idx, column=10).value
                ),
                "recharge_ham": {
                    "from_rainfall_monsoon": safe_float(
                        ws.cell(row=row_idx, column=11).value
                    ),
                    "from_other_sources_monsoon": safe_float(
                        ws.cell(row=row_idx, column=12).value
                    ),
                    "from_rainfall_non_monsoon": safe_float(
                        ws.cell(row=row_idx, column=13).value
                    ),
                    "from_other_sources_non_monsoon": safe_float(
                        ws.cell(row=row_idx, column=14).value
                    ),
                },
                "total_annual_recharge_ham": safe_float(
                    ws.cell(row=row_idx, column=15).value
                ),
                "total_natural_discharges_ham": safe_float(
                    ws.cell(row=row_idx, column=16).value
                ),
                "annual_extractable_resource_ham": safe_float(
                    ws.cell(row=row_idx, column=17).value
                ),
                "extraction_ham": {
                    "irrigation": safe_float(ws.cell(row=row_idx, column=18).value),
                    "industrial": safe_float(ws.cell(row=row_idx, column=19).value),
                    "domestic": safe_float(ws.cell(row=row_idx, column=20).value),
                    "total": safe_float(ws.cell(row=row_idx, column=21).value),
                },
                "annual_gw_allocation_domestic_2025_ham": safe_float(
                    ws.cell(row=row_idx, column=22).value
                ),
                "net_gw_availability_future_ham": safe_float(
                    ws.cell(row=row_idx, column=23).value
                ),
                "stage_of_extraction_percent": safe_float(
                    ws.cell(row=row_idx, column=24).value
                ),
                "categorization": safe_str(ws.cell(row=row_idx, column=25).value),
                "aquifer": safe_str(ws.cell(row=row_idx, column=26).value),
            }
            all_data["detailed_table"].append(record)

    wb.close()

    total_records = len(all_data["summary"]) + len(all_data["detailed_table"])
    print(
        f"    → Extracted {len(all_data['summary'])} summary + {len(all_data['detailed_table'])} detailed records"
    )

    return year, all_data


# =============================================================================
# MAIN ETL PIPELINE
# =============================================================================


def process_all_reports():
    """Process all Central, Annexure, and Attribute reports"""

    print("=" * 70)
    print("CENTRAL, ANNEXURE & ATTRIBUTE REPORTS ETL PIPELINE")
    print("=" * 70)
    print("\nField Naming Convention:")
    print("  • ham → hectare meters (1 ham = 10,000 cubic meters)")
    print("  • bcm → billion cubic meters")
    print("  • sq_km → square kilometers")
    print("  • ha → hectares")
    print("  • Categorization: safe, semi_critical, critical, over_exploited, saline")
    print()

    # Create output directories
    central_dir = OUTPUT_DIR / "central_reports"
    annexure_dir = OUTPUT_DIR / "annexures"
    attribute_dir = OUTPUT_DIR / "attribute_reports"

    central_dir.mkdir(parents=True, exist_ok=True)
    annexure_dir.mkdir(parents=True, exist_ok=True)
    attribute_dir.mkdir(parents=True, exist_ok=True)

    master_data = {
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "description": "Combined groundwater data from Central Reports, Annexures, and Attribute Reports",
            "units": {
                "ham": "hectare meters (1 ham = 10,000 cubic meters)",
                "bcm": "billion cubic meters",
                "sq_km": "square kilometers",
                "ha": "hectares",
                "mm": "millimeters",
            },
        },
        "central_reports": {},
        "annexure1_state_summary": {},
        "annexure2_district_resources": {},
        "annexure3_categorization": {},
        "annexure4_unit_details": {},
        "attribute_reports": {},
    }

    # Process Central Reports
    print("\n" + "-" * 50)
    print("PROCESSING CENTRAL REPORTS")
    print("-" * 50)
    central_files = sorted([f for f in DATA_DIR.glob("CentralReport*.xlsx")])
    print(f"Found {len(central_files)} central report files\n")

    for filepath in central_files:
        try:
            year, records = extract_central_report(filepath)
            master_data["central_reports"][year] = records

            # Save individual file
            output_file = central_dir / f"central_report_{year.replace('-', '_')}.json"
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(
                    {"year": year, "records": records}, f, indent=2, ensure_ascii=False
                )
        except Exception as e:
            print(f"  ERROR processing {filepath.name}: {e}")

    # Process Annexure 1
    print("\n" + "-" * 50)
    print("PROCESSING ANNEXURE 1 (State Summary)")
    print("-" * 50)
    annex1_files = sorted([f for f in DATA_DIR.glob("Annexure1*.xlsx")])
    print(f"Found {len(annex1_files)} Annexure 1 files\n")

    for filepath in annex1_files:
        try:
            year, records = extract_annexure1(filepath)
            master_data["annexure1_state_summary"][year] = records

            output_file = annexure_dir / f"annexure1_{year.replace('-', '_')}.json"
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(
                    {"year": year, "records": records}, f, indent=2, ensure_ascii=False
                )
        except Exception as e:
            print(f"  ERROR processing {filepath.name}: {e}")

    # Process Annexure 2
    print("\n" + "-" * 50)
    print("PROCESSING ANNEXURE 2 (District Resources)")
    print("-" * 50)
    annex2_files = sorted([f for f in DATA_DIR.glob("Annexure2*.xlsx")])
    print(f"Found {len(annex2_files)} Annexure 2 files\n")

    for filepath in annex2_files:
        try:
            year, records = extract_annexure2(filepath)
            master_data["annexure2_district_resources"][year] = records

            output_file = annexure_dir / f"annexure2_{year.replace('-', '_')}.json"
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(
                    {"year": year, "records": records}, f, indent=2, ensure_ascii=False
                )
        except Exception as e:
            print(f"  ERROR processing {filepath.name}: {e}")

    # Process Annexure 3
    print("\n" + "-" * 50)
    print("PROCESSING ANNEXURE 3 (Categorization - 6 Sheets)")
    print("-" * 50)
    annex3_files = sorted([f for f in DATA_DIR.glob("Annexure3*.xlsx")])
    print(f"Found {len(annex3_files)} Annexure 3 files\n")

    for filepath in annex3_files:
        try:
            year, data = extract_annexure3(filepath)
            master_data["annexure3_categorization"][year] = data

            output_file = annexure_dir / f"annexure3_{year.replace('-', '_')}.json"
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"  ERROR processing {filepath.name}: {e}")

    # Process Annexure 4
    print("\n" + "-" * 50)
    print("PROCESSING ANNEXURE 4 (Unit Details & Quality)")
    print("-" * 50)
    annex4_files = sorted([f for f in DATA_DIR.glob("Annexure4*.xlsx")])
    print(f"Found {len(annex4_files)} Annexure 4 files\n")

    for filepath in annex4_files:
        try:
            year, data = extract_annexure4(filepath)
            master_data["annexure4_unit_details"][year] = data

            output_file = annexure_dir / f"annexure4_{year.replace('-', '_')}.json"
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"  ERROR processing {filepath.name}: {e}")

    # Process Attribute Reports
    print("\n" + "-" * 50)
    print("PROCESSING ATTRIBUTE REPORTS")
    print("-" * 50)
    attr_files = sorted([f for f in DATA_DIR.glob("attributeReport*.xlsx")])
    print(f"Found {len(attr_files)} Attribute Report files\n")

    for filepath in attr_files:
        try:
            year, data = extract_attribute_report(filepath)
            master_data["attribute_reports"][year] = data

            output_file = (
                attribute_dir / f"attribute_report_{year.replace('-', '_')}.json"
            )
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"  ERROR processing {filepath.name}: {e}")

    # Save master JSON
    print("\n" + "-" * 50)
    print("SAVING MASTER JSON")
    print("-" * 50)

    master_file = OUTPUT_DIR / "all_reports_master.json"
    with open(master_file, "w", encoding="utf-8") as f:
        json.dump(master_data, f, indent=2, ensure_ascii=False)

    # Print summary
    print("\n" + "=" * 70)
    print("ETL COMPLETE - SUMMARY")
    print("=" * 70)

    total_central = sum(len(v) for v in master_data["central_reports"].values())
    total_annex1 = sum(len(v) for v in master_data["annexure1_state_summary"].values())
    total_annex2 = sum(
        len(v) for v in master_data["annexure2_district_resources"].values()
    )
    total_annex3 = sum(
        sum(len(sheet) for sheet in v.values() if isinstance(sheet, list))
        for v in master_data["annexure3_categorization"].values()
    )
    total_annex4 = sum(
        sum(len(sheet) for sheet in v.values() if isinstance(sheet, list))
        for v in master_data["annexure4_unit_details"].values()
    )
    total_attr = sum(
        len(v.get("summary", [])) + len(v.get("detailed_table", []))
        for v in master_data["attribute_reports"].values()
    )

    print(f"\nCentral Reports:      {total_central:,} records")
    print(f"Annexure 1:           {total_annex1:,} records")
    print(f"Annexure 2:           {total_annex2:,} records")
    print(f"Annexure 3:           {total_annex3:,} records")
    print(f"Annexure 4:           {total_annex4:,} records")
    print(f"Attribute Reports:    {total_attr:,} records")
    print(f"{'─' * 40}")
    print(
        f"TOTAL:                {total_central + total_annex1 + total_annex2 + total_annex3 + total_annex4 + total_attr:,} records"
    )

    print(f"\nOutput Files:")
    print(f"  • {central_dir}")
    print(f"  • {annexure_dir}")
    print(f"  • {attribute_dir}")
    print(f"  • {master_file}")

    return master_data


if __name__ == "__main__":
    process_all_reports()
